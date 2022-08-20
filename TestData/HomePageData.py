import openpyxl
from utilities.logger import Logger

class HomePageData:


    test_home_data=[{"firstname":"Rahul","lastname":"shetty","gender":"Male"},{"firstname":"Ankhita","lastname":"shetty","gender":"Female"}]

    @staticmethod
    def gettestData(test_case_name):
        log = Logger().get_logger()
        try:
            book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\PythonDemo.xlsx")
            Dict = {}
            sheet = book.active
            for i in range(1, sheet.max_row + 1):

                if sheet.cell(row=i, column=1).value == test_case_name:
                    for j in range(2, sheet.max_column + 1):
                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            return[Dict]
        except Exception as e:
            log.error("An Exception occurred:" + str(e))
            raise Exception(e)


